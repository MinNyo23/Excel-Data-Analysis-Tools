import json
import sys
from openpyxl import load_workbook

workbook = load_workbook(sys.argv[1], read_only=True, data_only=True)
result = {"sheetNames": workbook.sheetnames, "rows": {}}
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    result["rows"][sheet_name] = [list(row) for row in sheet.iter_rows(values_only=True)]
print(json.dumps(result, ensure_ascii=False, default=str))
