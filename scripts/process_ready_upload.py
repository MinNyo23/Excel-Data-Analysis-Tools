import base64, io, json, sys
import pandas as pd
from openpyxl import load_workbook
RENAME={'Employee Full Name':'Name','Employee ID':'Employee Registration Number','Mobile No':'Mobile Number','NRC No':'IdentityNumber','Father Name':'Contact Person'}
BLANK=['Title','Mobile Country Code','Marital Status','Identity Type','Contact Number','Country','State','City','Township']
ORDER=['Title','Name','Email','Employee Registration Number','Mobile Country Code','Mobile Number','Date of Birth','Gender','Marital Status','Identity Type','IdentityNumber','Contact Person','Contact Number','Country','Nationality','State','City','Township','Address']
def preview(d): return {'columns':[str(c) for c in d.columns],'rows':[[None if pd.isna(v) else (v.strftime('%m/%d/%Y') if isinstance(v,pd.Timestamp) else v) for v in r] for r in d.head(50).itertuples(index=False,name=None)]}
def go(p):
 d=pd.read_excel(io.BytesIO(base64.b64decode(p['file']['data'])),dtype=str).rename(columns=RENAME)
 if 'Date of Birth' in d:d['Date of Birth']=pd.to_datetime(d['Date of Birth'],format='mixed',dayfirst=True,errors='coerce')
 for c in BLANK:d[c]=''
 d=d[[c for c in ORDER if c in d]];o=io.BytesIO()
 with pd.ExcelWriter(o,engine='openpyxl') as w:d.to_excel(w,index=False)
 o.seek(0);wb=load_workbook(o);ws=wb.active
 if 'Date of Birth' in list(d.columns):
  i=list(d.columns).index('Date of Birth')+1
  for r in range(2,ws.max_row+1):ws.cell(r,i).number_format='mm/dd/yyyy'
 out=io.BytesIO();wb.save(out);return {'outputFilename':'Transformed_Employee_Data.xlsx','rowCount':len(d),'columnCount':len(d.columns),'preview':preview(d),'workbookBase64':base64.b64encode(out.getvalue()).decode()}
json.dump(go(json.load(sys.stdin)),sys.stdout,ensure_ascii=False,allow_nan=False)
