import base64
import io
import json
import sys

from openpyxl import load_workbook


MAX_GENERATED_WORKBOOK_BYTES = 40 * 1024 * 1024
FORMULA_PREFIXES = ("=", "+", "-", "@")


def needs_literal_text(value):
    return isinstance(value, str) and value.lstrip().startswith(FORMULA_PREFIXES)


def sanitize_workbook(workbook_base64):
    try:
        workbook_bytes = base64.b64decode(workbook_base64, validate=True)
    except (ValueError, TypeError) as error:
        raise ValueError("Invalid generated workbook payload.") from error
    if not workbook_bytes or len(workbook_bytes) > MAX_GENERATED_WORKBOOK_BYTES:
        raise ValueError("Generated workbook exceeds the safe download limit.")

    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False, keep_links=False)
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if needs_literal_text(cell.value):
                    # Excel treats a leading apostrophe as a literal-value marker;
                    # it is not displayed to the person opening the workbook.
                    cell.value = "'" + cell.value

    output = io.BytesIO()
    workbook.save(output)
    sanitized_bytes = output.getvalue()
    if not sanitized_bytes or len(sanitized_bytes) > MAX_GENERATED_WORKBOOK_BYTES:
        raise ValueError("Generated workbook exceeds the safe download limit.")
    return base64.b64encode(sanitized_bytes).decode("ascii")


payload = json.load(sys.stdin)
json.dump({"workbookBase64": sanitize_workbook(payload["workbookBase64"])}, sys.stdout)
